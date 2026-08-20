"""Staff import service — parse Excel/CSV штатное расписание, create users + positions.

Stage 1d of employee onboarding epic.

Flow:
1. HR uploads file (.xls, .xlsx or .csv) at /admin/staff
2. Backend parses, normalizes columns (Russian/English), returns preview
3. HR reviews preview (new users, matched users, new departments, new positions)
4. HR commits, backend creates/updates rows in transaction

CSV columns (case-insensitive, Russian OR English):
- personnel_number (required) - табельный номер
- first_name (required)
- last_name (required)
- department (required)
- position (required)
- email (optional)
- phone (optional)
- hire_date (optional, ISO format preferred)

Logic:
- User matched by normalized personnel_number within tenant
- Department is resolved/created before Position
- Position is resolved by normalized name inside that Department
- If user exists: update the imported profile and assign the resolved position
- If new: create with status='inactive', is_active=true (HR-managed)
  (password_hash=NULL - no self-service login)
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any
from uuid import UUID, uuid4

import xlrd  # type: ignore[import-untyped]
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.department import Department
from app.models.users import User
from app.modules.positions.models import Position

logger = logging.getLogger(__name__)


def _normalize_staff_text(value: Any) -> str:
    """Trim and collapse whitespace while preserving display casing."""
    return re.sub(r"\s+", " ", str(value or "").strip())


def normalize_staff_lookup(value: Any) -> str:
    """Return the shared case-insensitive key used by staff hierarchy flows."""
    return _normalize_staff_text(value).casefold()


def _normalized_row_values(row: Any) -> dict[str, str | date | None]:
    raw_hire_date = _normalize_staff_text(getattr(row, "hire_date", ""))
    branch = _normalize_staff_text(getattr(row, "branch", ""))
    return {
        "personnel_number": _normalize_staff_text(getattr(row, "personnel_number", "")),
        "first_name": _normalize_staff_text(getattr(row, "first_name", "")),
        "last_name": _normalize_staff_text(getattr(row, "last_name", "")),
        "department": _normalize_staff_text(getattr(row, "department", "")) or branch,
        "branch": branch,
        "position": _normalize_staff_text(getattr(row, "position", "")),
        "email": _normalize_staff_text(getattr(row, "email", "")).lower() or None,
        "phone": _normalize_staff_text(getattr(row, "phone", "")) or None,
        "hire_date": date.fromisoformat(raw_hire_date) if raw_hire_date else None,
    }


async def _load_staff_indexes(
    db: AsyncSession,
    tenant_id: UUID,
) -> tuple[
    dict[str, User],
    dict[str, list[User]],
    dict[str, Department],
    dict[UUID, Department],
    dict[tuple[str, str], Position],
]:
    """Load tenant-scoped users and hierarchy, including legacy text fallback."""
    users_result = await db.execute(
        select(User).where(
            User.tenant_id == tenant_id,
        )
    )
    users_by_pn: dict[str, User] = {}
    users_by_email: dict[str, list[User]] = {}
    for user in users_result.scalars().all():
        if user.tenant_id != tenant_id:
            continue
        key = normalize_staff_lookup(user.personnel_number)
        if key:
            users_by_pn.setdefault(key, user)
        email_key = normalize_staff_lookup(user.email)
        if email_key:
            users_by_email.setdefault(email_key, []).append(user)

    departments_result = await db.execute(select(Department).where(Department.tenant_id == tenant_id))
    departments_by_slug: dict[str, Department] = {}
    departments_by_id: dict[UUID, Department] = {}
    for department in departments_result.scalars().all():
        if department.tenant_id != tenant_id:
            continue
        for key in (
            normalize_staff_lookup(department.slug),
            normalize_staff_lookup(department.name),
        ):
            if key:
                departments_by_slug.setdefault(key, department)
        departments_by_id[department.id] = department

    positions_result = await db.execute(select(Position).where(Position.tenant_id == tenant_id))
    positions_by_key: dict[tuple[str, str], Position] = {}
    for position in positions_result.scalars().all():
        if position.tenant_id != tenant_id:
            continue
        position_key = normalize_staff_lookup(position.name)
        if not position_key:
            continue

        department = departments_by_id.get(position.department_id)
        if department is not None:
            positions_by_key.setdefault((str(department.id), position_key), position)

        # Legacy rows may have only Position.department. Keep this fallback
        # for reads and use it to backfill the canonical FK when written.
        legacy_department_key = normalize_staff_lookup(position.department)
        if legacy_department_key:
            positions_by_key.setdefault((legacy_department_key, position_key), position)

    return (
        users_by_pn,
        users_by_email,
        departments_by_slug,
        departments_by_id,
        positions_by_key,
    )


class StaffEmailConflictError(ValueError):
    """Raised when one tenant email would identify two staff records."""

    def __init__(self, email: str):
        self.email = email
        super().__init__(f"Email уже используется другим сотрудником: {email}")


def _assert_unique_staff_emails(
    parsed: ParsedFile,
    users_by_pn: dict[str, User],
    users_by_email: dict[str, list[User]],
) -> None:
    planned_email_owner: dict[str, str] = {}
    for row in parsed.rows:
        values = _normalized_row_values(row)
        email = normalize_staff_lookup(values["email"])
        if not email:
            continue
        pn = normalize_staff_lookup(values["personnel_number"])
        existing_by_pn = users_by_pn.get(pn)
        existing_email_owners = users_by_email.get(email, [])
        if len(existing_email_owners) > 1 or any(
            owner.id != getattr(existing_by_pn, "id", None) for owner in existing_email_owners
        ):
            raise StaffEmailConflictError(email)
        previous_pn = planned_email_owner.get(email)
        if previous_pn is not None and previous_pn != pn:
            raise StaffEmailConflictError(email)
        planned_email_owner[email] = pn


def _find_position(
    positions_by_key: dict[tuple[str, str], Position],
    department_key: str,
    department: Department | None,
    position_key: str,
) -> Position | None:
    if department is not None:
        position = positions_by_key.get((str(department.id), position_key))
        if position is not None:
            return position
    return positions_by_key.get((department_key, position_key))


@dataclass
class _ProjectedStaffUser:
    """Preview state after applying the preceding row for one personnel number."""

    first_name: str
    last_name: str
    email: str | None
    phone: str | None
    hire_date: date | None
    position_ref: object
    existing_user_id: str | None


# ── Column mapping ──────────────────────────────────────────────


# Maps recognized column headers (lowercase, trimmed) to canonical field names.
# Supports both Russian (HR's typical export) and English (template).
COLUMN_ALIASES: dict[str, str] = {
    # personnel_number
    "personnel_number": "personnel_number",
    "personnelnumber": "personnel_number",
    "табельный_номер": "personnel_number",
    "табельный номер": "personnel_number",
    "таб_номер": "personnel_number",
    "табельный№": "personnel_number",
    "employee_id": "personnel_number",
    "employeeid": "personnel_number",
    "tab_no": "personnel_number",
    "tabno": "personnel_number",
    # first_name
    "first_name": "first_name",
    "firstname": "first_name",
    "имя": "first_name",
    # last_name
    "last_name": "last_name",
    "lastname": "last_name",
    "фамилия": "last_name",
    # full_name
    "сотрудник": "full_name",
    "фио": "full_name",
    # department
    "department": "department",
    "отдел": "department",
    "подразделение": "department",
    "цех": "department",
    # branch (kept separate from legacy department)
    "branch": "branch",
    "branch_name": "branch",
    "филиал": "branch",
    "название филиала": "branch",
    "филиал атауы": "branch",
    # position
    "position": "position",
    "должность": "position",
    # email
    "email": "email",
    "e-mail": "email",
    "почта": "email",
    # phone
    "phone": "phone",
    "телефон": "phone",
    # hire_date
    "hire_date": "hire_date",
    "hiredate": "hire_date",
    "дата_приема": "hire_date",
    "дата приема": "hire_date",
}


REQUIRED_FIELDS = {"personnel_number", "first_name", "last_name", "department", "position"}
OPTIONAL_FIELDS = {"branch", "email", "phone", "hire_date", "full_name"}
ALL_FIELDS = REQUIRED_FIELDS | OPTIONAL_FIELDS


# ── Parsed row types ─────────────────────────────────────────────


@dataclass
class ParsedRow:
    """One row of the import file, normalized."""

    row_number: int  # 1-based row in original file (skipping header)
    personnel_number: str
    first_name: str
    last_name: str
    department: str
    position: str
    email: str | None = None
    phone: str | None = None
    hire_date: str | None = None  # ISO format if present
    branch: str = ""


@dataclass
class ParsedFile:
    """Result of parsing the uploaded file."""

    rows: list[ParsedRow]
    invalid_rows: list[dict]  # [{row_number, errors: [...], raw: {...}}]
    detected_columns: dict[str, str]  # original -> canonical
    missing_required_columns: list[str]
    total_rows_in_file: int
    raw_columns: list[str] = field(default_factory=list)
    sample_rows: list[dict[str, str]] = field(default_factory=list)
    suggested_mapping: dict[str, str] = field(default_factory=dict)
    sheet_name: str | None = None
    header_row: int = 1
    sheets: list[dict[str, Any]] = field(default_factory=list)


@dataclass
class PreviewItem:
    """One row in the preview (will_create / will_update / will_skip)."""

    row_number: int
    personnel_number: str
    first_name: str
    last_name: str
    department: str
    position: str
    email: str | None
    phone: str | None
    action: str  # 'create' | 'update' | 'skip'
    existing_user_id: str | None  # if action='update', set
    notes: list[str] = field(default_factory=list)


@dataclass
class PreviewResult:
    """Full preview returned to HR before commit."""

    items: list[PreviewItem]
    new_positions: list[str]  # (department, position) tuples that will be auto-created
    new_departments: list[str]  # departments that don't exist yet
    summary: dict[str, int]  # {'create': N, 'update': M, 'skip': K, 'new_positions': P}


# ── File parsing ────────────────────────────────────────────────


def _normalize_header(h: Any) -> str:
    """Lowercase + trim + collapse spaces."""
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _suggest_field_for_header(raw: str) -> str | None:
    """Best-effort match for messy HR exports."""
    normalized = _normalize_header(raw)
    direct = COLUMN_ALIASES.get(normalized)
    if direct:
        return direct
    tokens = set(re.findall(r"[a-zа-яё0-9]+", normalized))
    if "id" in tokens:
        return "personnel_number"
    if any(token == "тел" or token.startswith("телефон") for token in tokens):
        return "phone"
    compact = re.sub(r"[\s_\-№#./]+", "", normalized)
    rules: list[tuple[str, str]] = [
        ("personnel_number", "таб", ""),
        ("personnel_number", "таб", "номер"),
        ("personnel_number", "employee", "id"),
        ("personnel_number", "personnel", ""),
        ("first_name", "имя", ""),
        ("first_name", "employee", "name"),
        ("first_name", "name", "first"),
        ("full_name", "фио", ""),
        ("full_name", "fullname", ""),
        ("full_name", "full", "name"),
        ("last_name", "фам", ""),
        ("last_name", "family", ""),
        ("last_name", "surname", ""),
        ("last_name", "last", "name"),
        ("branch", "филиал", ""),
        ("branch", "branch", ""),
        ("department", "отдел", ""),
        ("department", "департамент", ""),
        ("department", "подраздел", ""),
        ("department", "department", ""),
        ("department", "division", ""),
        ("position", "долж", ""),
        ("position", "пози", ""),
        ("position", "job", "title"),
        ("position", "role", ""),
        ("position", "position", ""),
        ("email", "email", ""),
        ("email", "mail", ""),
        ("phone", "phone", ""),
        ("hire_date", "при", "дат"),
        ("hire_date", "hire", "date"),
    ]
    for canonical_field, a, b in rules:
        if a in compact and (not b or b in compact):
            return canonical_field
    return None


def _build_column_map(raw_columns: list[str], mapping: dict[str, str] | None = None) -> dict[str, str]:
    """Return raw_header -> canonical field mapping."""
    manual = {field: raw for field, raw in (mapping or {}).items() if raw}
    column_map: dict[str, str] = {}
    used_fields: set[str] = set()

    for canonical_field, raw in manual.items():
        if canonical_field in ALL_FIELDS and raw in raw_columns and canonical_field not in used_fields:
            column_map[raw] = canonical_field
            used_fields.add(canonical_field)

    for raw in raw_columns:
        if raw in column_map:
            continue
        canonical = _suggest_field_for_header(raw)
        if canonical and canonical not in used_fields:
            column_map[raw] = canonical
            used_fields.add(canonical)

    return column_map


def _suggested_mapping_from_column_map(column_map: dict[str, str]) -> dict[str, str]:
    return {canonical: raw for raw, canonical in column_map.items()}


def _split_full_name(full_name: str) -> tuple[str, str]:
    parts = [p for p in str(full_name).strip().split() if p]
    if len(parts) >= 2:
        # User has no dedicated patronymic column. Preserve every name part by
        # storing "Имя Отчество" in first_name and the leading surname in last_name.
        return " ".join(parts[1:]), parts[0]
    if len(parts) == 1:
        return parts[0], parts[0]
    return "", ""


def _missing_required_fields(column_map: dict[str, str]) -> set[str]:
    values = set(column_map.values())
    missing = REQUIRED_FIELDS - values
    if "full_name" in values:
        missing.discard("first_name")
        missing.discard("last_name")
    if "branch" in values:
        # A separate branch column is a valid structural parent for a direct
        # position; old files continue to require their department column.
        missing.discard("department")
    return missing


def _normalize_fields(fields: dict[str, str]) -> dict[str, str]:
    if fields.get("full_name") and (not fields.get("first_name") or not fields.get("last_name")):
        first, last = _split_full_name(fields["full_name"])
        fields.setdefault("first_name", first)
        fields.setdefault("last_name", last)
    return fields


def _sheet_score(sheet_name: str, raw_columns: list[str], sample_rows: list[dict[str, str]]) -> int:
    column_map = _build_column_map(raw_columns)
    values = set(column_map.values())
    score = (
        len(
            values
            & {"personnel_number", "first_name", "last_name", "full_name", "branch", "department", "position", "email"}
        )
        * 10
    )
    score += len(sample_rows)
    normalized_sheet = _normalize_header(sheet_name)
    if "сотруд" in normalized_sheet or "employee" in normalized_sheet:
        score += 80
    if "отдел" in normalized_sheet or "department" in normalized_sheet:
        score -= 40
    if "долж" in normalized_sheet or "position" in normalized_sheet:
        score -= 30
    if "personnel_number" in values:
        score += 70
    else:
        score -= 80
    if "email" in values:
        score += 10
    if {"personnel_number", "department", "position"} <= values and (
        {"first_name", "last_name"} <= values or "full_name" in values
    ):
        score += 100
    return score


def _xlsx_sheet_candidates(wb) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        max_row = min(ws.max_row or 0, 20)
        best: dict[str, Any] | None = None
        for header_row in range(1, max_row + 1):
            header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
            if not header_cells:
                continue
            raw_columns = [str(c).strip() if c is not None else "" for c in header_cells]
            if not any(raw_columns):
                continue
            data_rows = list(
                ws.iter_rows(
                    min_row=header_row + 1, max_row=min(ws.max_row or header_row, header_row + 5), values_only=True
                )
            )
            sample_rows = [
                {
                    raw_header: str(cell).strip() if cell is not None else ""
                    for raw_header, cell in zip(raw_columns, row, strict=False)
                }
                for row in data_rows
                if any(cell is not None and str(cell).strip() for cell in row)
            ]
            score = _sheet_score(ws.title, raw_columns, sample_rows)
            column_map = _build_column_map(raw_columns)
            missing_required = sorted(_missing_required_fields(column_map))
            normalized_sheet = _normalize_header(ws.title)
            is_reference_sheet = any(
                marker in normalized_sheet for marker in ("отдел", "department", "долж", "position")
            )
            sheet_kind = "employees" if not missing_required else "reference" if is_reference_sheet else "needs_mapping"
            if best is None or score > best["score"]:
                best = {
                    "sheet_name": ws.title,
                    "header_row": header_row,
                    "score": score,
                    "raw_columns": raw_columns,
                    "sample_rows": sample_rows,
                    "suggested_mapping": _suggested_mapping_from_column_map(column_map),
                    "missing_required_columns": missing_required,
                    "is_importable": not missing_required,
                    "sheet_kind": sheet_kind,
                }
        if best:
            candidates.append(best)
    return sorted(candidates, key=lambda c: c["score"], reverse=True)


def _parse_hire_date(s: str | None) -> str | None:
    """Try to parse hire_date. Returns ISO format or None."""
    if not s:
        return None
    s = str(s).strip()
    if not s:
        return None
    # Already ISO
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return s[:10]
    # DD.MM.YYYY or DD/MM/YYYY (Russian/European)
    m = re.match(r"^(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y if int(y) < 50 else "19" + y
        try:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            return None
    return None  # can't parse


def _canonical_branch_department(value: str) -> str:
    """Turn a legacy branch section label into a stable department name."""
    normalized = _normalize_staff_text(value)
    city_match = re.search(r"(?:^|\s)г\.\s*([^,(]+)", normalized, flags=re.IGNORECASE)
    if city_match:
        city = city_match.group(1).strip(" .")
        if city:
            return f"Филиал {city}"
    return normalized


def _is_branch_section(value: str) -> bool:
    normalized = _normalize_staff_text(value)
    return bool(
        re.search(r"(?:^|\s)г\.\s*", normalized, flags=re.IGNORECASE)
        or re.search(r"\bфилиал\b", normalized, flags=re.IGNORECASE)
    )


def _xls_cell_text(book: xlrd.book.Book, cell: xlrd.sheet.Cell) -> str:
    """Return a display-safe value while preserving common personnel-number formats."""
    if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, book.datemode).date().isoformat()
        except (OverflowError, TypeError, ValueError):
            return str(cell.value).strip()
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric = float(cell.value)
        if numeric.is_integer():
            # Personnel numbers are identifiers, not quantities.  Older HR
            # workbooks often store them as numeric cells with a display mask
            # such as ``0000``; preserve that mask instead of silently turning
            # employee ``0001`` into ``1`` during import.
            try:
                cell_format = book.xf_list[cell.xf_index]
                format_string = book.format_map[cell_format.format_key].format_str
                if re.fullmatch(r"0+", format_string):
                    return f"{int(numeric):0{len(format_string)}d}"
            except (AttributeError, IndexError, KeyError, TypeError):
                pass
            return str(int(numeric))
    return str(cell.value).strip()


def _xls_has_branch_sections(book: xlrd.book.Book, sheet: xlrd.sheet.Sheet, header_row_index: int) -> bool:
    for row_index in range(header_row_index + 1, sheet.nrows):
        values = [_xls_cell_text(book, sheet.cell(row_index, column)) for column in range(sheet.ncols)]
        populated = [value for value in values if value]
        if len(populated) == 1 and values and values[0] and _is_branch_section(values[0]):
            return True
    return False


def _xls_sheet_candidates(book: xlrd.book.Book) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for sheet in book.sheets():
        best: dict[str, Any] | None = None
        for header_row_index in range(min(sheet.nrows, 20)):
            raw_columns = [_xls_cell_text(book, sheet.cell(header_row_index, column)) for column in range(sheet.ncols)]
            if not any(raw_columns):
                continue
            sample_rows: list[dict[str, str]] = []
            for row_index in range(header_row_index + 1, min(sheet.nrows, header_row_index + 6)):
                values = [_xls_cell_text(book, sheet.cell(row_index, column)) for column in range(sheet.ncols)]
                if any(values):
                    sample_rows.append(dict(zip(raw_columns, values, strict=False)))
            column_map = _build_column_map(raw_columns)
            missing_required = _missing_required_fields(column_map)
            has_branch_sections = _xls_has_branch_sections(book, sheet, header_row_index)
            if has_branch_sections:
                missing_required.discard("department")
            candidate = {
                "sheet_name": sheet.name,
                "header_row": header_row_index + 1,
                "score": _sheet_score(sheet.name, raw_columns, sample_rows) + (40 if has_branch_sections else 0),
                "raw_columns": raw_columns,
                "sample_rows": sample_rows,
                "suggested_mapping": _suggested_mapping_from_column_map(column_map),
                "missing_required_columns": sorted(missing_required),
                "is_importable": not missing_required,
                "sheet_kind": "employees" if not missing_required else "needs_mapping",
                "has_branch_sections": has_branch_sections,
            }
            if best is None or candidate["score"] > best["score"]:
                best = candidate
        if best:
            candidates.append(best)
    return sorted(candidates, key=lambda candidate: candidate["score"], reverse=True)


def _xlsx_data_rows_with_merged_values(ws, *, min_row: int) -> list[tuple[Any, ...]]:
    """Return worksheet rows while expanding merged data-cell anchors.

    Tenant HR exports commonly merge a branch/department cell vertically for
    all employees in that group. OpenPyXL exposes only the top-left value and
    returns ``None`` for the remaining rows. Expanding only data rows preserves
    the workbook's meaning without inventing a generic fill-down rule for
    genuinely empty cells.
    """

    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col).value
        if anchor is None:
            continue
        for row_index in range(max(min_row, merged_range.min_row), merged_range.max_row + 1):
            for column_index in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_index, column_index)] = anchor

    rows: list[tuple[Any, ...]] = []
    for row_index, row in enumerate(
        ws.iter_rows(min_row=min_row, values_only=True),
        start=min_row,
    ):
        rows.append(
            tuple(
                merged_values.get((row_index, column_index), cell) if cell is None else cell
                for column_index, cell in enumerate(row, start=1)
            )
        )
    return rows


def parse_xls(
    content: bytes,
    mapping: dict[str, str] | None = None,
    sheet_name: str | None = None,
) -> ParsedFile:
    """Parse Excel 97-2003 .xls files, including branch section marker rows."""
    try:
        book = xlrd.open_workbook(file_contents=content, on_demand=True)
    except xlrd.biffh.XLRDError as exc:
        raise ValueError("Не удалось прочитать старый Excel-файл .xls.") from exc

    candidates = _xls_sheet_candidates(book)
    selected = next((candidate for candidate in candidates if candidate["sheet_name"] == sheet_name), None)
    if selected is None and candidates:
        selected = candidates[0]
    if selected is None:
        book.release_resources()
        return ParsedFile(
            rows=[],
            invalid_rows=[],
            detected_columns={},
            missing_required_columns=sorted(REQUIRED_FIELDS),
            total_rows_in_file=0,
        )

    sheet = book.sheet_by_name(selected["sheet_name"])
    header_row_index = int(selected["header_row"]) - 1
    raw_columns = [_xls_cell_text(book, sheet.cell(header_row_index, column)) for column in range(sheet.ncols)]
    column_map = _build_column_map(raw_columns, mapping)
    missing = _missing_required_fields(column_map)
    has_branch_sections = _xls_has_branch_sections(book, sheet, header_row_index)
    if has_branch_sections:
        missing.discard("department")
    if missing:
        book.release_resources()
        return ParsedFile(
            rows=[],
            invalid_rows=[],
            detected_columns=column_map,
            missing_required_columns=sorted(missing),
            total_rows_in_file=0,
            raw_columns=raw_columns,
            sample_rows=selected["sample_rows"],
            suggested_mapping=_suggested_mapping_from_column_map(column_map),
            sheet_name=sheet.name,
            header_row=header_row_index + 1,
            sheets=candidates,
        )

    rows: list[ParsedRow] = []
    invalid: list[dict[str, Any]] = []
    seen_personnel_numbers: set[str] = set()
    current_department = ""
    current_branch = ""
    for row_index in range(header_row_index + 1, sheet.nrows):
        values = [_xls_cell_text(book, sheet.cell(row_index, column)) for column in range(sheet.ncols)]
        if not any(values):
            continue
        populated = [value for value in values if value]
        if len(populated) == 1 and values[0] and _is_branch_section(values[0]):
            current_department = _canonical_branch_department(values[0])
            continue

        fields: dict[str, str] = {}
        raw_dict: dict[str, str] = {}
        for raw_header, value in zip(raw_columns, values, strict=False):
            canonical = column_map.get(raw_header)
            if canonical is None:
                raw_dict[raw_header] = value
            else:
                fields[canonical] = value
        if "branch" in column_map.values():
            if fields.get("branch"):
                current_branch = fields["branch"]
            elif current_branch:
                fields["branch"] = current_branch
        if not fields.get("department") and current_department:
            fields["department"] = current_department
        if fields.get("email"):
            fields["email"] = re.sub(r"\s+", "", fields["email"])
        fields = _normalize_fields(fields)

        errors = [
            f"Поле «{required}» пустое"
            for required in REQUIRED_FIELDS
            if not fields.get(required) and not (required == "department" and fields.get("branch"))
        ]
        personnel_number = fields.get("personnel_number", "").strip()
        if personnel_number:
            personnel_key = personnel_number.casefold()
            if personnel_key in seen_personnel_numbers:
                errors.append(f"Дубликат табельного номера «{personnel_number}»")
            seen_personnel_numbers.add(personnel_key)
        if errors:
            invalid.append({"row_number": row_index + 1, "errors": errors, "raw": raw_dict})
            continue

        rows.append(
            ParsedRow(
                row_number=row_index + 1,
                personnel_number=personnel_number,
                first_name=fields.get("first_name", "").strip(),
                last_name=fields.get("last_name", "").strip(),
                department=fields.get("department", "").strip(),
                position=fields.get("position", "").strip(),
                email=fields.get("email") or None,
                phone=fields.get("phone") or None,
                hire_date=_parse_hire_date(fields.get("hire_date")),
                branch=fields.get("branch", "").strip(),
            )
        )

    book.release_resources()
    return ParsedFile(
        rows=rows,
        invalid_rows=invalid,
        detected_columns=column_map,
        missing_required_columns=[],
        total_rows_in_file=len(rows) + len(invalid),
        raw_columns=raw_columns,
        sample_rows=selected["sample_rows"],
        suggested_mapping=_suggested_mapping_from_column_map(column_map),
        sheet_name=sheet.name,
        header_row=header_row_index + 1,
        sheets=candidates,
    )


def parse_csv(content: bytes, mapping: dict[str, str] | None = None, sheet_name: str | None = None) -> ParsedFile:
    """Parse CSV with Russian/English headers, return ParsedFile."""
    # Decode (try utf-8-sig first for BOM, then cp1251 for old Russian Excel exports)
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = content.decode("cp1251")
        except UnicodeDecodeError:
            text = content.decode("utf-8", errors="replace")

    reader = csv.DictReader(io.StringIO(text))
    raw_columns = [c for c in (reader.fieldnames or [])]
    raw_rows = list(reader)
    sample_rows = [{str(k): str(v).strip() if v is not None else "" for k, v in row.items()} for row in raw_rows[:5]]
    column_map = _build_column_map(raw_columns, mapping)

    missing = _missing_required_fields(column_map)
    if missing:
        return ParsedFile(
            rows=[],
            invalid_rows=[],
            detected_columns=column_map,
            missing_required_columns=sorted(missing),
            total_rows_in_file=0,
            raw_columns=raw_columns,
            sample_rows=sample_rows,
            suggested_mapping=_suggested_mapping_from_column_map(column_map),
            sheet_name=None,
            header_row=1,
            sheets=[],
        )

    rows: list[ParsedRow] = []
    invalid: list[dict] = []
    seen_pn: set[str] = set()
    current_branch = ""
    for i, raw_row in enumerate(raw_rows, start=2):  # row 1 = header
        # Skip empty rows
        if not any(v and str(v).strip() for v in raw_row.values()):
            continue

        # Extract fields
        fields: dict[str, str] = {}
        for raw, canonical in column_map.items():
            v = raw_row.get(raw)
            if v is None:
                continue
            fields[canonical] = str(v).strip()
        if "branch" in column_map.values():
            if fields.get("branch"):
                current_branch = fields["branch"]
            elif current_branch:
                fields["branch"] = current_branch
        fields = _normalize_fields(fields)

        # Validate required
        errors: list[str] = []
        for req in REQUIRED_FIELDS:
            if not fields.get(req) and not (req == "department" and fields.get("branch")):
                errors.append(f"Поле «{req}» пустое")

        pn = fields.get("personnel_number", "").strip()
        if pn:
            pn_norm = pn.lower()
            if pn_norm in seen_pn:
                errors.append(f"Дубликат табельного номера «{pn}»")
            seen_pn.add(pn_norm)

        if errors:
            invalid.append(
                {
                    "row_number": i,
                    "errors": errors,
                    "raw": dict(raw_row),
                }
            )
            continue

        rows.append(
            ParsedRow(
                row_number=i,
                personnel_number=pn,
                first_name=fields.get("first_name", "").strip(),
                last_name=fields.get("last_name", "").strip(),
                department=fields.get("department", "").strip(),
                position=fields.get("position", "").strip(),
                email=fields.get("email") or None,
                phone=fields.get("phone") or None,
                hire_date=_parse_hire_date(fields.get("hire_date")),
                branch=fields.get("branch", "").strip(),
            )
        )

    return ParsedFile(
        rows=rows,
        invalid_rows=invalid,
        detected_columns=column_map,
        missing_required_columns=[],
        total_rows_in_file=len(rows) + len(invalid),
        raw_columns=raw_columns,
        sample_rows=sample_rows,
        suggested_mapping=_suggested_mapping_from_column_map(column_map),
        sheet_name=None,
        header_row=1,
        sheets=[],
    )


def parse_xlsx(content: bytes, mapping: dict[str, str] | None = None, sheet_name: str | None = None) -> ParsedFile:
    """Parse Excel .xlsx via openpyxl. Returns ParsedFile (same shape as parse_csv)."""
    wb = load_workbook(
        io.BytesIO(content),
        read_only=False,
        data_only=True,
        keep_links=False,
    )
    sheet_candidates = _xlsx_sheet_candidates(wb)
    selected_sheet = None
    if sheet_name:
        selected_sheet = next((c for c in sheet_candidates if c["sheet_name"] == sheet_name), None)
    if selected_sheet is None and sheet_candidates:
        selected_sheet = sheet_candidates[0]
    ws = wb[selected_sheet["sheet_name"]] if selected_sheet else wb.active
    header_row = int(selected_sheet["header_row"]) if selected_sheet else 1

    # Header row
    header_cells = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True), None)
    if not header_cells:
        return ParsedFile(
            rows=[],
            invalid_rows=[],
            detected_columns={},
            missing_required_columns=list(REQUIRED_FIELDS),
            total_rows_in_file=0,
            sheets=sheet_candidates,
            sheet_name=ws.title,
            header_row=header_row,
        )

    raw_columns = [str(c).strip() if c is not None else "" for c in header_cells]
    raw_data_rows = _xlsx_data_rows_with_merged_values(
        ws,
        min_row=header_row + 1,
    )
    sample_rows = [
        {
            raw_header: str(cell).strip() if cell is not None else ""
            for raw_header, cell in zip(raw_columns, row, strict=False)
        }
        for row in raw_data_rows[:5]
    ]
    column_map = _build_column_map(raw_columns, mapping)

    missing = _missing_required_fields(column_map)
    if missing:
        wb.close()
        return ParsedFile(
            rows=[],
            invalid_rows=[],
            detected_columns=column_map,
            missing_required_columns=sorted(missing),
            total_rows_in_file=0,
            raw_columns=raw_columns,
            sample_rows=sample_rows,
            suggested_mapping=_suggested_mapping_from_column_map(column_map),
            sheet_name=ws.title,
            header_row=header_row,
            sheets=sheet_candidates,
        )

    rows: list[ParsedRow] = []
    invalid: list[dict] = []
    seen_pn: set[str] = set()
    current_branch = ""
    for i, row in enumerate(raw_data_rows, start=header_row + 1):
        # Skip empty rows
        if not any(v is not None and str(v).strip() for v in row):
            continue

        # Map cells to fields
        fields: dict[str, str] = {}
        raw_dict: dict[str, Any] = {}
        for raw_header, cell in zip(raw_columns, row, strict=False):
            canonical = column_map.get(raw_header)
            if canonical is None:
                # Unmapped column - keep as raw for invalid rows
                raw_dict[raw_header] = str(cell).strip() if cell is not None else ""
            else:
                v = str(cell).strip() if cell is not None else ""
                fields[canonical] = v
        if "branch" in column_map.values():
            if fields.get("branch"):
                current_branch = fields["branch"]
            elif current_branch:
                fields["branch"] = current_branch
        fields = _normalize_fields(fields)

        errors: list[str] = []
        for req in REQUIRED_FIELDS:
            if not fields.get(req) and not (req == "department" and fields.get("branch")):
                errors.append(f"Поле «{req}» пустое")

        pn = fields.get("personnel_number", "").strip()
        if pn:
            pn_norm = pn.lower()
            if pn_norm in seen_pn:
                errors.append(f"Дубликат табельного номера «{pn}»")
            seen_pn.add(pn_norm)

        if errors:
            invalid.append(
                {
                    "row_number": i,
                    "errors": errors,
                    "raw": raw_dict,
                }
            )
            continue

        rows.append(
            ParsedRow(
                row_number=i,
                personnel_number=pn,
                first_name=fields.get("first_name", "").strip(),
                last_name=fields.get("last_name", "").strip(),
                department=fields.get("department", "").strip(),
                position=fields.get("position", "").strip(),
                email=fields.get("email") or None,
                phone=fields.get("phone") or None,
                hire_date=_parse_hire_date(fields.get("hire_date")),
                branch=fields.get("branch", "").strip(),
            )
        )

    wb.close()
    return ParsedFile(
        rows=rows,
        invalid_rows=invalid,
        detected_columns=column_map,
        missing_required_columns=[],
        total_rows_in_file=len(rows) + len(invalid),
        raw_columns=raw_columns,
        sample_rows=sample_rows,
        suggested_mapping=_suggested_mapping_from_column_map(column_map),
        sheet_name=ws.title,
        header_row=header_row,
        sheets=sheet_candidates,
    )


def parse_upload(
    filename: str,
    content: bytes,
    mapping: dict[str, str] | None = None,
    sheet_name: str | None = None,
) -> ParsedFile:
    """Dispatch based on file extension."""
    name = filename.lower()
    if name.endswith(".csv"):
        return parse_csv(content, mapping=mapping, sheet_name=sheet_name)
    if name.endswith(".xlsx"):
        return parse_xlsx(content, mapping=mapping, sheet_name=sheet_name)
    if name.endswith(".xls"):
        return parse_xls(content, mapping=mapping, sheet_name=sheet_name)
    raise ValueError(f"Формат файла не поддерживается: {filename}. Используйте .xls, .xlsx или .csv.")


# ── Preview (against current DB) ─────────────────────────────────


async def build_preview(
    db: AsyncSession,
    tenant_id: UUID,
    parsed: ParsedFile,
) -> PreviewResult:
    """Match parsed rows against existing users/positions/departments, return preview."""

    (
        users_by_pn,
        users_by_email,
        departments_by_slug,
        _,
        positions_by_key,
    ) = await _load_staff_indexes(db, tenant_id)
    _assert_unique_staff_emails(parsed, users_by_pn, users_by_email)

    items: list[PreviewItem] = []
    new_positions: dict[tuple[str, str], tuple[str, str]] = {}
    new_departments: dict[str, str] = {}
    projected_users: dict[str, _ProjectedStaffUser] = {}

    for row in parsed.rows:
        values = _normalized_row_values(row)
        pn_norm = normalize_staff_lookup(values["personnel_number"])
        existing = users_by_pn.get(pn_norm)

        department_key = normalize_staff_lookup(values["department"])
        department = departments_by_slug.get(department_key)
        position_key = normalize_staff_lookup(values["position"])
        position = _find_position(positions_by_key, department_key, department, position_key)
        if department is None and department_key:
            new_departments.setdefault(department_key, values["department"] or "")
        if position is None and department_key and position_key:
            new_positions.setdefault(
                (department_key, position_key),
                (values["department"] or "", values["position"] or ""),
            )

        position_ref: object = position.id if position is not None else ("new", department_key, position_key)
        projected = projected_users.get(pn_norm)
        is_repeated_row = projected is not None
        if projected is None and existing is not None:
            projected = _ProjectedStaffUser(
                first_name=(existing.first_name or "").strip(),
                last_name=(existing.last_name or "").strip(),
                email=(existing.email or "").strip().lower() or None,
                phone=existing.phone,
                hire_date=existing.hire_date,
                position_ref=existing.position_id,
                existing_user_id=str(existing.id),
            )

        notes: list[str] = []
        if projected is not None:
            # Missing optional columns mean "leave the stored value unchanged".
            # A regular import must not erase employment data simply because a
            # narrower spreadsheet was uploaded later.
            if values["email"] is None:
                values["email"] = projected.email
            if values["phone"] is None:
                values["phone"] = projected.phone
            if values["hire_date"] is None:
                values["hire_date"] = projected.hire_date
            if projected.first_name != values["first_name"]:
                notes.append(f"имя: «{projected.first_name}» → «{values['first_name']}»")
            if projected.last_name != values["last_name"]:
                notes.append(f"фамилия: «{projected.last_name}» → «{values['last_name']}»")
            if projected.email != values["email"]:
                notes.append(f"email: «{projected.email or '—'}» → «{values['email'] or '—'}»")
            if projected.phone != values["phone"]:
                notes.append(f"телефон: «{projected.phone or '—'}» → «{values['phone'] or '—'}»")
            if projected.hire_date != values["hire_date"]:
                notes.append(f"дата приёма: «{projected.hire_date or '—'}» → «{values['hire_date'] or '—'}»")
            if projected.position_ref != position_ref:
                notes.append(f"новая должность: «{values['position']}» (отдел «{values['department']}»)")

            items.append(
                PreviewItem(
                    row_number=row.row_number,
                    personnel_number=values["personnel_number"] or "",
                    first_name=values["first_name"] or "",
                    last_name=values["last_name"] or "",
                    department=values["department"] or "",
                    position=values["position"] or "",
                    email=values["email"],
                    phone=values["phone"],
                    action="update" if notes else "skip",
                    existing_user_id=projected.existing_user_id,
                    notes=notes or ["Повторный ряд: без изменений" if is_repeated_row else "Без изменений"],
                )
            )
        else:
            if position is None:
                notes.append(f"новая должность: «{values['position']}» в «{values['department']}»")
            items.append(
                PreviewItem(
                    row_number=row.row_number,
                    personnel_number=values["personnel_number"] or "",
                    first_name=values["first_name"] or "",
                    last_name=values["last_name"] or "",
                    department=values["department"] or "",
                    position=values["position"] or "",
                    email=values["email"],
                    phone=values["phone"],
                    action="create",
                    existing_user_id=None,
                    notes=notes,
                )
            )

        projected_users[pn_norm] = _ProjectedStaffUser(
            first_name=values["first_name"] or "",
            last_name=values["last_name"] or "",
            email=values["email"],
            phone=values["phone"],
            hire_date=values["hire_date"],
            position_ref=position_ref,
            existing_user_id=projected.existing_user_id if projected else None,
        )

    summary = {
        "create": sum(1 for i in items if i.action == "create"),
        "update": sum(1 for i in items if i.action == "update"),
        "skip": sum(1 for i in items if i.action == "skip"),
        "new_positions": len(new_positions),
        "new_departments": len(new_departments),
    }
    return PreviewResult(
        items=items,
        new_positions=sorted([f"{dept} / {pos}" for dept, pos in new_positions.values()]),
        new_departments=sorted(new_departments.values()),
        summary=summary,
    )


# ── Commit (apply changes) ───────────────────────────────────────


async def commit_import(
    db: AsyncSession,
    tenant_id: UUID,
    parsed: ParsedFile,
    *,
    commit_changes: bool = True,
    apply_rules: bool = True,
) -> dict:
    """Apply the import: create/update users + create new positions.

    Returns {created: N, updated: M, skipped: K, positions_created: P}.
    """
    (
        users_by_pn,
        users_by_email,
        departments_by_slug,
        departments_by_id,
        positions_by_key,
    ) = await _load_staff_indexes(db, tenant_id)
    _assert_unique_staff_emails(parsed, users_by_pn, users_by_email)

    created = 0
    updated = 0
    skipped = 0
    positions_created = 0
    # Track users that need rule recomputation after the hierarchy write.
    # The recomputation itself runs inline after the import commit below.
    affected_user_ids: list[UUID] = []

    for row in parsed.rows:
        values = _normalized_row_values(row)
        pn_norm = normalize_staff_lookup(values["personnel_number"])
        existing = users_by_pn.get(pn_norm)

        # Resolve/create the canonical Department before Position.
        department_key = normalize_staff_lookup(values["department"])
        department = departments_by_slug.get(department_key)
        if department is None:
            department = Department(
                id=uuid4(),
                tenant_id=tenant_id,
                name=values["department"] or "",
                slug=department_key,
                description="",
            )
            db.add(department)
            await db.flush()
            departments_by_slug[department_key] = department
            departments_by_id[department.id] = department

        position_key = normalize_staff_lookup(values["position"])
        pos = _find_position(positions_by_key, department_key, department, position_key)
        if pos is None:
            pos = Position(
                id=uuid4(),
                tenant_id=tenant_id,
                name=values["position"] or "",
                department=department.name,
                level="",
                responsibilities="",
                requirements="",
                employee_count=0,
                department_id=department.id,
            )
            db.add(pos)
            await db.flush()
            positions_by_key[(str(department.id), position_key)] = pos
            positions_by_key.setdefault((department_key, position_key), pos)
            positions_created += 1
        else:
            # Legacy positions retain their text column for compatibility,
            # but all rows touched by a new write get the canonical FK.
            pos.department_id = department.id
            pos.department = department.name
            positions_by_key[(str(department.id), position_key)] = pos
            positions_by_key.setdefault((department_key, position_key), pos)

        if existing:
            # Keep previously stored optional values when the import omits
            # their columns or leaves the cells blank.
            if values["email"] is None:
                values["email"] = existing.email
            if values["phone"] is None:
                values["phone"] = existing.phone
            if values["hire_date"] is None:
                values["hire_date"] = existing.hire_date
            # Check if anything actually changes
            changed = False
            if (existing.first_name or "").strip() != values["first_name"]:
                existing.first_name = values["first_name"]
                changed = True
            if (existing.last_name or "").strip() != values["last_name"]:
                existing.last_name = values["last_name"]
                changed = True
            if (existing.email or "").strip().lower() != (values["email"] or ""):
                existing.email = values["email"]
                changed = True
            if (existing.phone or "") != (values["phone"] or ""):
                existing.phone = values["phone"]
                changed = True
            if existing.hire_date != values["hire_date"]:
                existing.hire_date = values["hire_date"]
                changed = True
            # Position changed — that's also a trigger for apply-rules
            position_changed = False
            if existing.position_id != pos.id:
                existing.position_id = pos.id
                existing.is_active = True
                changed = True
                position_changed = True
            if changed:
                updated += 1
                # Recompute only if the user's position actually
                # changed — name/email updates don't move rules.
                if position_changed or not existing.position_id:
                    affected_user_ids.append(existing.id)
            else:
                skipped += 1
        else:
            user = User(
                id=uuid4(),
                tenant_id=tenant_id,
                personnel_number=values["personnel_number"],
                email=values["email"],
                phone=values["phone"],
                hire_date=values["hire_date"],
                first_name=values["first_name"] or "",
                last_name=values["last_name"] or "",
                role="student",  # bulk import always creates students; HR promotes separately
                is_active=True,
                position_id=pos.id,
                password_hash=None,  # no self-service login - HR-managed
                status="active",
            )
            db.add(user)
            await db.flush()
            created += 1
            affected_user_ids.append(user.id)
            # Invalidate cache so next row can find this user (duplicate-PN check)
            users_by_pn[pn_norm] = user

    if commit_changes:
        await db.commit()
    else:
        await db.flush()

    # P0-1 (TZ §2.6): trigger apply-rules inline so the import
    # is actually useful. Pre-fix the router dispatched to Celery,
    # but on Render free tier there's no worker process and the
    # task silently dropped — new staff had no enrollments.
    # Inline asyncio + Redis progress is enough for v1.0 (a few
    # hundred users fits in <10s). The standalone
    # /admin/staff/apply-rules endpoint still exists for
    # retroactive retries (see staff_import_router).
    apply_rules_task_id: str | None = None
    if affected_user_ids and apply_rules:
        from app.core import redis_progress
        from app.modules.positions.batch_service import apply_rules_for_users

        apply_rules_task_id = redis_progress.new_task_id()
        await redis_progress.init_task(apply_rules_task_id, total=len(affected_user_ids))
        await redis_progress.mark_started(apply_rules_task_id)

        # Chunked per TZ §2.6 (50 users per chunk). The chunked
        # call is the same `apply_rules_for_users` that the
        # retroactive endpoint uses; the kernel handles the
        # recompute invariants regardless of chunk size.
        chunk_size = 50
        aggregate_added = 0
        aggregate_removed = 0
        aggregate_failed = 0
        for i in range(0, len(affected_user_ids), chunk_size):
            chunk = affected_user_ids[i : i + chunk_size]
            try:
                # Per TZ §2.6: 'успех импорта не зависит от
                # успеха apply' — but we now run inside the same
                # function, so we catch + log per-chunk to keep
                # the import summary clean.
                outcome = await apply_rules_for_users(db, chunk)
                aggregate_added += outcome.added
                aggregate_removed += outcome.removed
                # One Redis tick per processed user so the UI
                # progress bar advances smoothly.
                for _ in chunk:
                    await redis_progress.increment_done(
                        apply_rules_task_id,
                        added=0,
                        removed=0,
                    )
            except Exception as exc:  # noqa: BLE001 — apply-rules
                # failures must NOT roll back the import.
                aggregate_failed += len(chunk)
                logger.exception(
                    "apply-rules inline chunk failed (chunk_size=%d): %s",
                    len(chunk),
                    exc,
                )
                for _ in chunk:
                    await redis_progress.increment_failed(apply_rules_task_id)

        # Final state in Redis.
        result_payload = {
            "users_processed": len(affected_user_ids) - aggregate_failed,
            "added": aggregate_added,
            "removed": aggregate_removed,
            "failed_chunks": aggregate_failed // chunk_size if chunk_size else 0,
        }
        if aggregate_failed == 0:
            await redis_progress.mark_success(apply_rules_task_id, result_payload)
        elif aggregate_failed < len(affected_user_ids):
            # Partial — mark SUCCESS (we did what we could) but
            # the failed count is in the result payload so the
            # UI can surface a warning.
            await redis_progress.mark_success(apply_rules_task_id, result_payload)
        else:
            # Total failure — mark FAILURE so the UI shows red.
            await redis_progress.mark_failure(
                apply_rules_task_id,
                f"All {aggregate_failed} affected users failed apply-rules",
            )

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "positions_created": positions_created,
        "affected_user_ids": [str(uid) for uid in affected_user_ids],
        "apply_rules_task_id": apply_rules_task_id,
    }


async def create_manual_staff_member(
    db: AsyncSession,
    tenant_id: UUID,
    *,
    personnel_number: str,
    first_name: str,
    last_name: str,
    department: str,
    position: str,
    email: str | None = None,
    phone: str | None = None,
) -> dict:
    """Create one HR-managed learner without uploading a staff file."""
    row = ParsedRow(
        row_number=1,
        personnel_number=_normalize_staff_text(personnel_number),
        first_name=_normalize_staff_text(first_name),
        last_name=_normalize_staff_text(last_name),
        department=_normalize_staff_text(department),
        position=_normalize_staff_text(position),
        email=_normalize_staff_text(email).lower() or None,
        phone=_normalize_staff_text(phone) or None,
    )

    parsed = ParsedFile(
        rows=[row],
        invalid_rows=[],
        detected_columns={},
        missing_required_columns=[],
        total_rows_in_file=1,
    )
    return await commit_import(db, tenant_id, parsed)
