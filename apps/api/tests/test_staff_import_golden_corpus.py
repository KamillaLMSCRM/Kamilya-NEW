"""Golden corpus + spec-driven staff-import tests (parse + preview).

Covers T02 requested import formats:
- legacy sectioned XLS
- incremental / full branch coverage
- multi-sheet
- merged cells
- RU/KK/EN headers
- empty branches
- duplicate identifiers
- ambiguous name/email conflict
- rename with stable code
- move position
- partial and malformed workbook
"""

from __future__ import annotations

from io import BytesIO
from uuid import UUID, uuid4

import pytest
from openpyxl import Workbook

from app.models.department import Department
from app.models.users import User
from app.modules.positions.models import Position
from app.modules.users.staff_import_service import (
    StaffEmailConflictError,
    build_preview,
    parse_upload,
)

from .test_staff_import_xls import LEGACY_STAFF_XLS


def _xlsx_bytes(
    rows: list[dict[str, str]],
    headers: list[str],
    *,
    sheet_name: str = "Сотрудники",
    extra_sheets: dict[str, list[list[str]]] | None = None,
    merged_cell: str | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    for row_i, row in enumerate(rows, start=2):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row_i, column=col, value=row.get(header, ""))
    if merged_cell:
        ws.merge_cells(merged_cell)
    if extra_sheets:
        for title, lines in extra_sheets.items():
            sheet = wb.create_sheet(title=title)
            for r, values in enumerate(lines, start=1):
                for c, value in enumerate(values, start=1):
                    sheet.cell(row=r, column=c, value=value)
    output = BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()


def _make_db(seed: dict | None, tenant_id: UUID | None = None) -> _MemorySession:
    tenant_id = tenant_id or uuid4()
    session = _MemorySession(tenant_id=tenant_id)
    if not seed:
        return session

    for raw in seed.get("departments", []):
        department = Department(
            id=uuid4(),
            tenant_id=tenant_id,
            name=raw["name"],
            slug=raw["slug"],
        )
        session.departments.append(department)
        position = Position(
            id=uuid4(),
            tenant_id=tenant_id,
            name=raw["position"],
            department=raw["name"],
            department_id=department.id,
            level="",
            responsibilities="",
            requirements="",
            employee_count=0,
        )
        session.positions.append(position)
        session._positions_by_name[position.name] = position
        session._positions_by_department[department.name] = position

    for raw in seed.get("users", []):
        user = User(
            id=uuid4(),
            tenant_id=tenant_id,
            personnel_number=raw["personnel_number"],
            first_name=raw["first_name"],
            last_name=raw["last_name"],
            email=raw["email"],
            role="student",
            is_active=True,
            status="active",
        )
        position_ref = raw.get("position_ref")
        if isinstance(position_ref, str):
            normalized = position_ref.strip()
            user.position_id = session._positions_by_name.get(normalized) or session._positions_by_department.get(
                normalized
            )
            if user.position_id is not None:
                user.position_id = user.position_id.id
        else:
            if session.positions:
                user.position_id = session.positions[0].id
        session.users.append(user)
    return session


class _MemorySession:
    def __init__(self, tenant_id: UUID):
        self.tenant_id = tenant_id
        self.users: list[User] = []
        self.departments: list[Department] = []
        self.positions: list[Position] = []
        self._department_by_name: dict[str, Department] = {}
        self._positions_by_name: dict[str, Position] = {}
        self._positions_by_department: dict[str, Position] = {}

    async def execute(self, statement):
        entity = statement.column_descriptions[0]["entity"]
        if entity is User:
            return _ScalarResult(self.users)
        if entity is Department:
            return _ScalarResult(self.departments)
        if entity is Position:
            return _ScalarResult(self.positions)
        raise AssertionError(f"Unexpected entity: {entity}")

    def add(self, value):
        if isinstance(value, User):
            self.users.append(value)
        elif isinstance(value, Department):
            self.departments.append(value)
        elif isinstance(value, Position):
            self.positions.append(value)
        else:
            raise AssertionError(f"Unexpected object: {value!r}")

    async def flush(self):
        return None

    async def commit(self):
        return None


class _ScalarResult:
    def __init__(self, values):
        self._values = list(values)

    def scalars(self):
        return self

    def all(self):
        return self._values


GOLDEN_SPECS: list[dict] = [
    {
        "name": "legacy-sectioned-xls",
        "filename": "legacy-staff.xls",
        "bytes": lambda: LEGACY_STAFF_XLS,
        "parse": {"sheet_name": "Лист1", "header_row": 2, "rows": 4, "invalid_rows": 0},
    },
    {
        "name": "incremental-5-branches",
        "filename": "incremental-5-branches.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "P-101",
                    "ФИО": "Арайл Жансая",
                    "Должность": "Аналитик",
                    "Отдел": "Филиал Север 1",
                    "Email": "ara-101@tenant.example",
                },
                {
                    "Табельный номер": "P-102",
                    "ФИО": "Мейрам Жанай",
                    "Должность": "Младший аналитик",
                    "Отдел": "Филиал Север 2",
                    "Email": "meya-102@tenant.example",
                },
                {
                    "Табельный номер": "P-103",
                    "ФИО": "Алия Бектемир",
                    "Должность": "Руководитель",
                    "Отдел": "Филиал Восток 3",
                    "Email": "ali-103@tenant.example",
                },
                {
                    "Табельный номер": "P-104",
                    "ФИО": "Нурлан Арзына",
                    "Должность": "Координатор",
                    "Отдел": "Филиал Юг 4",
                    "Email": "nura-104@tenant.example",
                },
                {
                    "Табельный номер": "P-105",
                    "ФИО": "Айым Жолдас",
                    "Должность": "Специалист",
                    "Отдел": "Филиал Запад 5",
                    "Email": "aima-105@tenant.example",
                },
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 5, "invalid_rows": 0},
        "preview": {"summary": {"create": 5, "update": 0, "skip": 0, "new_positions": 5, "new_departments": 5}},
    },
    {
        "name": "full-7-branches",
        "filename": "full-7-branches.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "P-106",
                    "ФИО": "Санжар Толе",
                    "Должность": "Старший специалист",
                    "Отдел": "Филиал Центр 6",
                    "Email": "sanj-106@tenant.example",
                },
                {
                    "Табельный номер": "P-107",
                    "ФИО": "Ирина Нур",
                    "Должность": "Администратор",
                    "Отдел": "Филиал Экспресс 7",
                    "Email": "irina-107@tenant.example",
                },
            ]
            + [
                {
                    "Табельный номер": f"P-10{i}",
                    "ФИО": f"Заполнитель {i}",
                    "Должность": "Оператор",
                    "Отдел": f"Филиал Доп {i}",
                    "Email": f"add-{i}@tenant.example",
                }
                for i in range(1, 6)
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 7, "invalid_rows": 0},
        "preview": {"summary": {"create": 7, "update": 0, "skip": 0, "new_positions": 7, "new_departments": 7}},
    },
    {
        "name": "multi-sheet",
        "filename": "multi-sheet.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Employee ID": "EMP-201",
                    "Full Name": "Ariana Demo",
                    "Job Title": "Trainer",
                    "Department": "Sales",
                    "Email": "ariana-201@example.test",
                    "Phone": "+7 700 200 01",
                },
                {
                    "Employee ID": "EMP-202",
                    "Full Name": "Borys Test",
                    "Job Title": "Coach",
                    "Department": "Support",
                    "Email": "borys-202@example.test",
                    "Phone": "+7 700 200 02",
                },
            ],
            ["Employee ID", "Full Name", "Job Title", "Department", "Email", "Phone"],
            extra_sheets={
                "Отделы": [
                    ["Код", "Название отдела"],
                    ["DEP-01", "Sales"],
                    ["DEP-02", "Support"],
                ],
                "Метки": [["Примечание"], ["Данные для примера"]],
            },
        ),
        "parse": {
            "sheet_name": "Сотрудники",
            "rows": 2,
            "invalid_rows": 0,
            "sheet_candidates": ["Отделы", "Метки", "Сотрудники"],
        },
        "mapping": {
            "personnel_number": "Employee ID",
            "full_name": "Full Name",
            "position": "Job Title",
            "department": "Department",
            "email": "Email",
            "phone": "Phone",
        },
        "preview": {"summary": {"create": 2, "update": 0, "skip": 0, "new_positions": 2, "new_departments": 2}},
    },
    {
        "name": "merged-cells",
        "filename": "merged-cells.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-M01",
                    "ФИО": "Бахтияр Нур",
                    "Должность": "Аналитик",
                    "Отдел": "Филиал М1",
                    "Email": "m1@example.test",
                },
                {
                    "Табельный номер": "EMP-M02",
                    "ФИО": "Динара Төлеген",
                    "Должность": "Аналитик",
                    "Отдел": "Филиал М2",
                    "Email": "m2@example.test",
                },
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
            merged_cell="D2:D3",
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 2, "invalid_rows": 0},
        "preview": {"summary": {"create": 2, "update": 0, "skip": 0, "new_positions": 1, "new_departments": 1}},
    },
    {
        "name": "headers-ru",
        "filename": "headers-ru.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-RU1",
                    "Фамилия": "Жумаш",
                    "Имя": "Арай",
                    "Должность": "Аналитик",
                    "Отдел": "Регион RU",
                    "Email": "ru-1@example.test",
                }
            ],
            ["Табельный номер", "Фамилия", "Имя", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 0},
        "preview": {"summary": {"create": 1, "update": 0, "skip": 0, "new_positions": 1, "new_departments": 1}},
    },
    {
        "name": "headers-kk",
        "filename": "headers-kk.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Қызметкер коды": "EMP-KK1",
                    "Тегі": "Жанбота",
                    "Аты": "Серу",
                    "Лауазым": "Аналитик",
                    "Бөлім": "Регион KK",
                    "Поштасы": "kk-1@example.test",
                }
            ],
            ["Қызметкер коды", "Тегі", "Аты", "Лауазым", "Бөлім", "Поштасы", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 0},
        "mapping": {
            "personnel_number": "Қызметкер коды",
            "last_name": "Тегі",
            "first_name": "Аты",
            "position": "Лауазым",
            "department": "Бөлім",
            "email": "Поштасы",
            "phone": "Телефон",
        },
        "preview": {"summary": {"create": 1, "update": 0, "skip": 0, "new_positions": 1, "new_departments": 1}},
    },
    {
        "name": "headers-en",
        "filename": "headers-en.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Employee ID": "EMP-EN1",
                    "Last Name": "Harris",
                    "First Name": "Ari",
                    "Position": "Analyst",
                    "Department": "Region EN",
                    "Email": "en-1@example.test",
                }
            ],
            ["Employee ID", "Last Name", "First Name", "Position", "Department", "Email", "Phone"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 0},
        "mapping": {
            "personnel_number": "Employee ID",
            "first_name": "First Name",
            "last_name": "Last Name",
            "position": "Position",
            "department": "Department",
            "email": "Email",
            "phone": "Phone",
        },
        "preview": {"summary": {"create": 1, "update": 0, "skip": 0, "new_positions": 1, "new_departments": 1}},
    },
    {
        "name": "empty-branches",
        "filename": "empty-branches.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-E01",
                    "ФИО": "Лайла Нур",
                    "Должность": "Бренд",
                    "Отдел": "Филиал Главный",
                    "Email": "empty-1@example.test",
                }
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
            extra_sheets={"Пустые филиалы": [["Филиал"], ["Филиал без сотрудников 1"], ["Филиал без сотрудников 2"]]},
        ),
        "parse": {
            "sheet_name": "Сотрудники",
            "rows": 1,
            "invalid_rows": 0,
            "sheets_include": ["Пустые филиалы"],
        },
        "preview": {"summary": {"create": 1, "update": 0, "skip": 0, "new_positions": 1, "new_departments": 1}},
    },
    {
        "name": "duplicate-identifiers",
        "filename": "duplicate-identifiers.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-DUP",
                    "ФИО": "Анна Ану",
                    "Должность": "Оператор",
                    "Отдел": "Служба",
                    "Email": "dup-1@example.test",
                },
                {
                    "Табельный номер": "EMP-DUP",
                    "ФИО": "Анна Ану",
                    "Должность": "Оператор",
                    "Отдел": "Служба",
                    "Email": "dup-2@example.test",
                },
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 1},
    },
    {
        "name": "ambiguous-names",
        "filename": "ambiguous-names.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-AM1",
                    "ФИО": "Рауан Бек",
                    "Должность": "Спец",
                    "Отдел": "HR",
                    "Email": "shared@example.test",
                },
                {
                    "Табельный номер": "EMP-AM2",
                    "ФИО": "Рауан Бек",
                    "Должность": "Спец",
                    "Отдел": "HR",
                    "Email": "shared@example.test",
                },
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 2, "invalid_rows": 0},
        "preview_expect_error": StaffEmailConflictError,
    },
    {
        "name": "rename-stable-code",
        "filename": "rename-stable-code.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-RS1",
                    "ФИО": "Мира Алия",
                    "Должность": "Старший специалист",
                    "Отдел": "Отдел операций",
                    "Email": "rename@example.test",
                }
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 0},
        "seed": {
            "departments": [
                {"name": "Отдел операций", "slug": "otdel-operatsiy", "position": "Старший специалист"},
            ],
            "users": [
                {
                    "personnel_number": "EMP-RS1",
                    "first_name": "Мира",
                    "last_name": "Старое",
                    "email": "rename@example.test",
                    "position_ref": "Старший специалист",
                },
            ],
        },
        "preview": {"summary": {"create": 0, "update": 1, "skip": 0, "new_positions": 0, "new_departments": 0}},
        "seed_expected": {"expected_action": "update", "expected_note_contains": ["имя", "фамилия"]},
    },
    {
        "name": "move-position",
        "filename": "move-position.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-MV1",
                    "ФИО": "Талгат Кай",
                    "Должность": "Старший наставник",
                    "Отдел": "Отдел обучения",
                    "Email": "move@example.test",
                }
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 1, "invalid_rows": 0},
        "seed": {
            "departments": [
                {"name": "Отдел обучения", "slug": "otdel-obucheniya", "position": "Тьютор"},
            ],
            "users": [
                {
                    "personnel_number": "EMP-MV1",
                    "first_name": "Талгат",
                    "last_name": "Кай",
                    "email": "move@example.test",
                    "position_ref": "Тьютор",
                },
            ],
        },
        "preview": {"summary": {"create": 0, "update": 1, "skip": 0, "new_positions": 1, "new_departments": 0}},
        "seed_expected": {"expected_action": "update", "expected_note_contains": ["новая должность"]},
    },
    {
        "name": "partial-workbook",
        "filename": "partial-workbook.xlsx",
        "bytes": lambda: _xlsx_bytes(
            [
                {
                    "Табельный номер": "EMP-P01",
                    "ФИО": "Елдара Аян",
                    "Должность": "Курьер",
                    "Отдел": "Операции",
                    "Email": "part-01@example.test",
                },
                {
                    "Табельный номер": "EMP-P02",
                    "ФИО": "Саяхмет Гани",
                    "Должность": "Помощник",
                    "Отдел": "Операции",
                    "Email": "part-02@example.test",
                },
            ],
            ["Табельный номер", "ФИО", "Должность", "Отдел", "Email", "Телефон"],
        ),
        "parse": {"sheet_name": "Сотрудники", "rows": 2, "invalid_rows": 0},
        "preview": {"summary": {"create": 2, "update": 0, "skip": 0, "new_positions": 2, "new_departments": 1}},
    },
    {
        "name": "malformed-workbook",
        "filename": "malformed-workbook.xlsx",
        "bytes": lambda: b"not-an-excel-workbook",
        "parse_error": Exception,
    },
]


@pytest.mark.parametrize("case", GOLDEN_SPECS, ids=[case["name"] for case in GOLDEN_SPECS])
def test_golden_staff_import_parse(case: dict):
    filename = case["filename"]
    content = case["bytes"]()
    mapping = case.get("mapping")
    if "parse_error" in case:
        with pytest.raises(case["parse_error"]):
            parse_upload(filename, content, mapping=mapping)
        return

    parsed = parse_upload(filename, content, mapping=mapping)

    parse_exp = case["parse"]
    assert parsed.sheet_name == parse_exp["sheet_name"]
    assert len(parsed.rows) == parse_exp["rows"]
    assert len(parsed.invalid_rows) == parse_exp["invalid_rows"]
    if "header_row" in parse_exp:
        assert parsed.header_row == parse_exp["header_row"]
    if parse_exp.get("sheet_candidates"):
        assert sorted([sheet["sheet_name"] for sheet in parsed.sheets]) == sorted(parse_exp["sheet_candidates"])
    if parse_exp.get("sheets_include"):
        assert set(parse_exp["sheets_include"]).issubset({sheet["sheet_name"] for sheet in parsed.sheets})


@pytest.mark.asyncio
@pytest.mark.parametrize("case", GOLDEN_SPECS, ids=[case["name"] for case in GOLDEN_SPECS])
async def test_golden_staff_import_preview(case: dict):
    if "preview" not in case and "preview_expect_error" not in case:
        return

    parsed = parse_upload(case["filename"], case["bytes"](), mapping=case.get("mapping"))
    tenant_id = uuid4()
    db = _make_db(case.get("seed"), tenant_id=tenant_id)
    if "preview_expect_error" in case:
        with pytest.raises(case["preview_expect_error"]):
            await build_preview(db, tenant_id, parsed)
        return

    preview = await build_preview(db, tenant_id, parsed)
    expected = case["preview"]["summary"]
    assert preview.summary["create"] == expected["create"]
    assert preview.summary["update"] == expected["update"]
    assert preview.summary["skip"] == expected["skip"]
    assert len(preview.new_departments) == expected["new_departments"]
    assert len(preview.new_positions) == expected["new_positions"]

    seed_expected = case.get("seed_expected")
    if seed_expected:
        assert len(preview.items) == 1
        assert preview.items[0].action == seed_expected["expected_action"]
        notes = " ".join(preview.items[0].notes).lower()
        for needle in seed_expected["expected_note_contains"]:
            assert needle in notes
