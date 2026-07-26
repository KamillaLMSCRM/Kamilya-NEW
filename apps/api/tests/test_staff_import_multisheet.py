"""Regression coverage for multi-sheet staffing workbooks."""

from io import BytesIO

from openpyxl import Workbook

from app.modules.users.staff_import_service import parse_upload


def _workbook_bytes() -> bytes:
    workbook = Workbook()
    employees = workbook.active
    employees.title = "Сотрудники"
    employees.append(["Штатное расписание"])
    employees.append([])
    employees.append(
        [
            "Табельный №",
            "ФИО",
            "Должность",
            "Отдел",
            "Email",
            "Телефон",
        ]
    )
    employees.append(
        [
            "EMP-001",
            "Ахметов Айбек Жанатович",
            "Инженер",
            "IT",
            "aybek@example.test",
            "+7 700 000 00 01",
        ]
    )

    departments = workbook.create_sheet("Отделы")
    departments.append(["Код", "Название отдела", "Руководитель (ФИО)"])
    departments.append(["DEP-01", "IT", "Ахметов Айбек Жанатович"])

    positions = workbook.create_sheet("Должности")
    positions.append(
        [
            "Код должности",
            "Название должности",
            "Отдел",
            "Обязательные курсы",
        ]
    )
    positions.append(["POS-01", "Инженер", "IT", "Охрана труда"])

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def test_multisheet_parser_selects_only_importable_employee_sheet():
    parsed = parse_upload("staff.xlsx", _workbook_bytes())

    assert parsed.sheet_name == "Сотрудники"
    assert parsed.header_row == 3
    assert len(parsed.rows) == 1

    candidates = {sheet["sheet_name"]: sheet for sheet in parsed.sheets}
    assert candidates["Сотрудники"]["is_importable"] is True
    assert candidates["Сотрудники"]["missing_required_columns"] == []
    assert candidates["Отделы"]["is_importable"] is False
    assert candidates["Должности"]["is_importable"] is False
    assert "personnel_number" in candidates["Должности"]["missing_required_columns"]


def test_non_employee_sheet_does_not_infer_phone_from_required_courses():
    parsed = parse_upload(
        "staff.xlsx",
        _workbook_bytes(),
        sheet_name="Должности",
    )

    assert parsed.sheet_name == "Должности"
    assert "phone" not in parsed.suggested_mapping
    assert parsed.rows == []
    assert parsed.missing_required_columns
